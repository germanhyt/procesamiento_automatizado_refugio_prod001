# -*- coding: utf-8 -*-
"""
Detección de locatarios sin carga reciente en FileStore (cierre_caja pendientes).
"""
from __future__ import annotations

import csv
import logging
import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional, Set

import openpyxl
import xlrd

from app.core.constants import LOCATARIOS, MESES_ES
from app.services.file_store_service import (
    ZONA_LIMA,
    _dir_locatario_pendientes,
    get_semana_actual_lima,
    get_upload_base,
    list_pendientes_locatario,
    parse_fecha_desde_nombre_archivo,
)

logger = logging.getLogger(__name__)

MODOS_PERIODO_NOTIFICACIONES = frozenset({"ultima_semana", "semana_actual", "rango_libre", "ultimos_dias"})

_MAX_ROWS_SCAN = 200
_MAX_COLS_SCAN = 30
_MAX_ROWS_FECHA_COL = 3000
# Filas iniciales donde puede estar la cabecera real del detalle (no solo fila 1).
_MAX_HEADER_SEARCH_ROWS = 60


def _try_parse_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None

    value = value.strip()
    if not value:
        return None

    formatos = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d/%m/%y",
        "%Y/%m/%d",
    )
    for fmt in formatos:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _scan_xlsx(path: Path) -> Optional[date]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(
            max_row=_MAX_ROWS_SCAN,
            max_col=_MAX_COLS_SCAN,
            values_only=True,
        ):
            for cell_value in row:
                result = _try_parse_date(cell_value)
                if result is not None:
                    return result
    finally:
        wb.close()
    return None


def _scan_xls(path: Path) -> Optional[date]:
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    row_limit = min(ws.nrows, _MAX_ROWS_SCAN)
    col_limit = min(ws.ncols, _MAX_COLS_SCAN)

    for row_idx in range(row_limit):
        for col_idx in range(col_limit):
            cell = ws.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    tup = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                    return date(tup[0], tup[1], tup[2])
                except Exception:  # noqa: BLE001
                    pass
            if cell.ctype == xlrd.XL_CELL_TEXT:
                result = _try_parse_date(cell.value)
                if result is not None:
                    return result
    return None


def _scan_csv(path: Path) -> Optional[date]:
    for encoding in ("utf-8", "latin-1"):
        try:
            with path.open(encoding=encoding, errors="replace", newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)

                for row_idx, row in enumerate(reader):
                    if row_idx >= _MAX_ROWS_SCAN:
                        break
                    for cell_value in row[:_MAX_COLS_SCAN]:
                        result = _try_parse_date(cell_value)
                        if result is not None:
                            return result
            break
        except UnicodeDecodeError:
            continue
    return None


def _iter_dias_calendario(inicio: date, fin: date):
    d = inicio
    while d <= fin:
        yield d
        d += timedelta(days=1)


def _normalizar_header_fecha(h: Any) -> str:
    """Minúsculas, sin tildes, espacios colapsados (para comparar encabezados)."""
    if h is None:
        return ""
    s = str(h).replace("\ufeff", "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s).strip()
    return s


# Encabezados de metadatos del reporte (no son fecha operativa / detalle).
_FECHA_ENCABEZADO_EXCLUIDOS_PREFIJO = (
    "fecha de generacion",
    "fecha generacion",
)


def _encabezado_fecha_operativa_descartado(n: str) -> bool:
    if not n:
        return True
    for pref in _FECHA_ENCABEZADO_EXCLUIDOS_PREFIJO:
        if n == pref or n.startswith(pref + " "):
            return True
    return False


def _prioridad_columna_fecha_operativa(n: str) -> Optional[int]:
    """
    Menor número = más prioridad. None = no usar esta columna.

    No es un SQL LIKE genérico: cada regla compara el encabezado ya normalizado
    (_normalizar_header_fecha: minúsculas, sin tildes, espacios colapsados) con
    prefijos/igualdades concretos y exclusiones (p. ej. «fecha de generacion»).

    Orden: Fecha exacta → Fecha creación… → Fecha venta… → Fecha comprobante… → ^fecha(\\s|_)…
    """
    if _encabezado_fecha_operativa_descartado(n):
        return None
    if n == "fecha":
        return 0
    if n.startswith("fecha creacion"):
        return 1
    if n.startswith("fecha venta"):
        return 2
    if n.startswith("fecha comprobante"):
        return 3
    if re.match(r"^fecha(\s|_)", n):
        return 4
    return None


def _find_fecha_column_index(headers: tuple[Any, ...]) -> Optional[int]:
    best_i: Optional[int] = None
    best_p: Optional[int] = None
    for i, h in enumerate(headers):
        n = _normalizar_header_fecha(h)
        p = _prioridad_columna_fecha_operativa(n)
        if p is None:
            continue
        if best_p is None or p < best_p:
            best_p = p
            best_i = i
    return best_i


def _extract_fechas_bajo_encabezado(
    rows: list[tuple[Any, ...]],
    header_row_idx: int,
    col_idx: int,
) -> set[date]:
    """Lee fechas en rows[header_row_idx+1 : …] en la columna col_idx."""
    out: set[date] = set()
    end = min(len(rows), header_row_idx + 1 + _MAX_ROWS_FECHA_COL)
    for r in range(header_row_idx + 1, end):
        row = rows[r]
        if not row or col_idx >= len(row):
            continue
        parsed = _try_parse_date(row[col_idx])
        if parsed is not None:
            out.add(parsed)
    return out


def _buscar_fila_encabezado_fecha(rows: list[tuple[Any, ...]]) -> Optional[tuple[int, int]]:
    """
    Primera fila (entre 0.._MAX_HEADER_SEARCH_ROWS-1) que tenga una columna fecha reconocida
    y al menos una fecha parseable en filas siguientes.
    """
    search_limit = min(len(rows), _MAX_HEADER_SEARCH_ROWS)
    for i in range(search_limit):
        idx = _find_fecha_column_index(tuple(rows[i]))
        if idx is None:
            continue
        found = _extract_fechas_bajo_encabezado(rows, i, idx)
        if found:
            return (i, idx)
    return None


def _parse_celda_fecha_xlrd(cell: Any, wb: Any) -> Optional[date]:
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            tup = xlrd.xldate_as_tuple(cell.value, wb.datemode)
            return date(tup[0], tup[1], tup[2])
        except Exception:  # noqa: BLE001
            return None
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return _try_parse_date(cell.value)
    if cell.ctype == xlrd.XL_CELL_TEXT:
        return _try_parse_date(cell.value)
    return None


def _extract_fechas_bajo_encabezado_xls(
    ws: Any,
    wb: Any,
    header_row_idx: int,
    col_idx: int,
) -> set[date]:
    out: set[date] = set()
    row_limit = min(ws.nrows, header_row_idx + 1 + _MAX_ROWS_FECHA_COL)
    for r in range(header_row_idx + 1, row_limit):
        parsed = _parse_celda_fecha_xlrd(ws.cell(r, col_idx), wb)
        if parsed is not None:
            out.add(parsed)
    return out


def _extract_fechas_columna_xlsx(path: Path) -> set[date]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        cap = _MAX_HEADER_SEARCH_ROWS + _MAX_ROWS_FECHA_COL + 2
        rows: list[tuple[Any, ...]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(tuple(row) if row is not None else ())
            if i >= cap - 1:
                break
        hit = _buscar_fila_encabezado_fecha(rows)
        if hit is None:
            return set()
        h_idx, col_idx = hit
        return _extract_fechas_bajo_encabezado(rows, h_idx, col_idx)
    finally:
        wb.close()


def _extract_fechas_columna_xls(path: Path) -> set[date]:
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        return set()
    search_limit = min(ws.nrows, _MAX_HEADER_SEARCH_ROWS)
    for i in range(search_limit):
        header = tuple(ws.row_values(i))
        idx = _find_fecha_column_index(header)
        if idx is None:
            continue
        found = _extract_fechas_bajo_encabezado_xls(ws, wb, i, idx)
        if found:
            return found
    return set()


def _extract_fechas_columna_csv(path: Path) -> set[date]:
    for encoding in ("utf-8", "latin-1"):
        try:
            with path.open(encoding=encoding, errors="replace", newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                cap = _MAX_HEADER_SEARCH_ROWS + _MAX_ROWS_FECHA_COL + 2
                rows: list[tuple[Any, ...]] = []
                for i, row in enumerate(reader):
                    rows.append(tuple(row))
                    if i >= cap - 1:
                        break
                hit = _buscar_fila_encabezado_fecha(rows)
                if hit is None:
                    return set()
                h_idx, col_idx = hit
                return _extract_fechas_bajo_encabezado(rows, h_idx, col_idx)
        except UnicodeDecodeError:
            continue
    return set()


def collect_fechas_operacion_desde_archivo(path: Path, nombre: str) -> set[date]:
    """
    Fechas de operación: se busca la fila de cabecera del detalle en las primeras
    _MAX_HEADER_SEARCH_ROWS filas; columna por prioridad (Fecha, Fecha creación…,
    Fecha venta…, Fecha comprobante…, u otras «fecha » / «fecha_»; se excluye
    «Fecha de generación»). Si no hay tabla detectada, una fecha vía contenido/nombre/mtime.
    """
    ext = path.suffix.lower()
    col_dates: set[date] = set()
    try:
        if ext == ".xlsx":
            col_dates = _extract_fechas_columna_xlsx(path)
        elif ext == ".xls":
            col_dates = _extract_fechas_columna_xls(path)
        elif ext == ".csv":
            col_dates = _extract_fechas_columna_csv(path)
    except Exception as exc:  # noqa: BLE001
        logger.warning("collect_fechas_operacion: error en %s — %s", path.name, exc)
    if col_dates:
        return col_dates
    fecha_una, _ = _fecha_mas_reciente_del_archivo(path, nombre)
    if fecha_una is not None:
        return {fecha_una}
    return set()


def find_first_date_in_file(path: Path) -> Optional[date]:
    """
    Primera celda parseable como fecha (arriba→abajo, izquierda→derecha).
    Soporta .csv, .xlsx, .xls.
    """
    ext = path.suffix.lower()
    try:
        if ext == ".xlsx":
            return _scan_xlsx(path)
        if ext == ".xls":
            return _scan_xls(path)
        if ext == ".csv":
            return _scan_csv(path)
    except Exception as exc:  # noqa: BLE001
        logger.warning("find_first_date_in_file: error leyendo %s — %s", path.name, exc)
    return None


def _fecha_mas_reciente_del_archivo(path: Path, nombre: str) -> tuple[Optional[date], str]:
    fecha = find_first_date_in_file(path)
    if fecha is not None:
        return fecha, "contenido"

    fecha = parse_fecha_desde_nombre_archivo(nombre)
    if fecha is not None:
        return fecha, "nombre"

    try:
        mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=ZONA_LIMA)
        return mtime.date(), "mtime"
    except OSError:
        pass

    return None, "sin_fecha"


def _rango_ultima_semana_completa_lima() -> tuple[date, date, str]:
    """Semana ISO completa anterior a la semana calendario que contiene hoy (Lima)."""
    lunes_dt, domingo_dt, _, _ = get_semana_actual_lima()
    lunes_d = lunes_dt.date() if isinstance(lunes_dt, datetime) else lunes_dt
    lunes_prev = lunes_d - timedelta(days=7)
    domingo_prev = lunes_prev + timedelta(days=6)
    iso_week = lunes_prev.isocalendar()[1]
    mes_nombre = MESES_ES[lunes_prev.month - 1]
    nombre = f"semana{iso_week}_{lunes_prev.day}_{domingo_prev.day}_{mes_nombre}"
    return lunes_prev, domingo_prev, nombre


def _rango_semana_actual_lima() -> tuple[date, date, str]:
    lunes_dt, domingo_dt, nombre, _ = get_semana_actual_lima()
    ld = lunes_dt.date() if isinstance(lunes_dt, datetime) else lunes_dt
    dd = domingo_dt.date() if isinstance(domingo_dt, datetime) else domingo_dt
    return ld, dd, nombre


def resolver_periodo_notificaciones(
    modo: str,
    *,
    dias: Optional[int],
    fecha_inicio: Optional[date],
    fecha_fin: Optional[date],
) -> tuple[date, date, str, bool]:
    """
    Devuelve (inicio, fin, etiqueta, ventana_rodante).
    - ventana_rodante=True: cuenta carga si mejor_fecha >= inicio (sin techo; modo ultimos_dias).
    - ventana_rodante=False: inicio <= mejor_fecha <= fin.
    """
    modo_norm = (modo or "ultima_semana").strip().lower()
    hoy = datetime.now(ZONA_LIMA).date()

    if modo_norm == "ultima_semana":
        i, f, lbl = _rango_ultima_semana_completa_lima()
        return i, f, lbl, False
    if modo_norm == "semana_actual":
        i, f, lbl = _rango_semana_actual_lima()
        return i, f, lbl, False
    if modo_norm == "rango_libre":
        if fecha_inicio is None or fecha_fin is None:
            raise ValueError("rango_libre requiere fecha_inicio y fecha_fin (YYYY-MM-DD)")
        i, f = fecha_inicio, fecha_fin
        if i > f:
            i, f = f, i
        return i, f, f"rango_{i}_{f}", False
    if modo_norm == "ultimos_dias":
        d = dias if dias is not None and dias > 0 else 7
        inicio = hoy - timedelta(days=d)
        return inicio, hoy, f"ultimos_{d}_dias", True
    raise ValueError(
        "modo debe ser ultima_semana | semana_actual | rango_libre | ultimos_dias",
    )

# Subida de pendientes: https://datarefugio.gcbprojects.site/fuentes
def _sugerencia_notificacion(nombre_loc: str, faltantes: list[date]) -> str:
    if not faltantes:
        return ""
    partes = ", ".join(f"{d.day:02d}/{d.month:02d}/{d.year}" for d in faltantes)
    url_carga = "https://datarefugio.gcbprojects.site/fuentes"
    if len(faltantes) == 1:
        enunciado_fechas = f"el día {partes}"
    else:
        enunciado_fechas = f"los días {partes}"
    return (
        "Buenos días.\n\n"
        f"Les informamos que, según nuestro registro en Refugio Data, el local {nombre_loc} "
        f"tiene pendiente la carga de reportes de cierre de caja correspondiente a {enunciado_fechas}.\n\n"
        "Les agradecemos completar la subida en la plataforma cuando les sea posible:\n"
        f"{url_carga}\n\n"
        "Quedamos atentos ante cualquier consulta.\n\n"
        "Saludos cordiales."
    )


def evaluar_locatarios_pendientes_periodo(
    periodo_inicio: date,
    periodo_fin: date,
    *,
    ventana_rodante: bool = False,
) -> list[dict[str, Any]]:
    """
    Cada día calendario en [periodo_inicio, periodo_fin] debe tener al menos una fecha de
    operación detectada en los pendientes (columna Fecha o, si no existe, una fecha por archivo).

    ventana_rodante se ignora para la grilla: el periodo ya viene acotado al resolver.
    """
    _ = ventana_rodante
    hoy = datetime.now(ZONA_LIMA).date()
    base = get_upload_base()
    resultado: list[dict[str, Any]] = []

    dias_esperados_list = list(_iter_dias_calendario(periodo_inicio, periodo_fin))
    dias_esperados_set: Set[date] = set(dias_esperados_list)

    for loc in LOCATARIOS:
        codigo = loc["codigo"]
        nombres_archivo = list_pendientes_locatario(codigo)

        cobertura_total: set[date] = set()

        for nombre in nombres_archivo:
            path = _dir_locatario_pendientes(base, codigo) / nombre
            if not path.is_file():
                continue
            cobertura_total |= collect_fechas_operacion_desde_archivo(path, nombre)

        registrados_en_periodo = cobertura_total & dias_esperados_set
        faltantes_en_periodo = sorted(dias_esperados_set - registrados_en_periodo)

        mejor_en_periodo = max(registrados_en_periodo) if registrados_en_periodo else None
        mejor_global = max(cobertura_total) if cobertura_total else None

        ultimo_upload = (
            mejor_en_periodo.isoformat()
            if mejor_en_periodo is not None
            else (mejor_global.isoformat() if mejor_global is not None else None)
        )

        ref_antiguedad = mejor_global if mejor_global is not None else mejor_en_periodo
        dias_sin_subir: Optional[int]
        if ref_antiguedad is not None:
            dias_sin_subir = (hoy - ref_antiguedad).days
        else:
            dias_sin_subir = None

        alerta = len(faltantes_en_periodo) > 0
        fuente_fecha = "sin_archivos" if not cobertura_total else "dias_en_archivos"

        sugerencia = (
            _sugerencia_notificacion(loc["name"], faltantes_en_periodo)
            if faltantes_en_periodo
            else None
        )

        resultado.append({
            "codigo": codigo,
            "nombre": loc["name"],
            "ultimo_upload": ultimo_upload,
            "dias_sin_subir": dias_sin_subir,
            "alerta": alerta,
            "fuente_fecha": fuente_fecha,
            "dias_con_registro": [d.isoformat() for d in sorted(registrados_en_periodo)],
            "dias_faltantes": [d.isoformat() for d in faltantes_en_periodo],
            "sugerencia_notificacion": sugerencia,
        })

        logger.debug(
            "[notificaciones] %s → registro_periodo=%s faltantes=%s alerta=%s",
            codigo,
            len(registrados_en_periodo),
            len(faltantes_en_periodo),
            alerta,
        )

    return resultado


def lista_dias_periodo_iso(periodo_inicio: date, periodo_fin: date) -> list[str]:
    """Días calendario inclusivos en ISO (orden cronológico)."""
    return [d.isoformat() for d in _iter_dias_calendario(periodo_inicio, periodo_fin)]


def get_locatarios_sin_upload_semanal(dias: int = 7) -> list[dict[str, Any]]:
    """Compatibilidad: últimos N días en ventana rodante (mismo criterio que antes)."""
    hoy = datetime.now(ZONA_LIMA).date()
    start = hoy - timedelta(days=dias)
    return evaluar_locatarios_pendientes_periodo(start, hoy, ventana_rodante=True)


def nombre_semana_actual_lima() -> str:
    _, _, nombre, _ = get_semana_actual_lima()
    return nombre
