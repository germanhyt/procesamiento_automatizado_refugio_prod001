# -*- coding: utf-8 -*-
"""
Diagnóstico FileStore: A06_DON_MELCHOR — encabezado primera fila vs detección real
(fila de cabecera + columna por prioridad Fecha / Fecha creación / …).

  pytest tests/test_a06_don_melchor_fecha_diagnostico.py -s -v

Sin archivos en pendientes → skip.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
import pytest
import xlrd

from app.services import notificaciones_service as ns
from app.services.file_store_service import _dir_locatario_pendientes, get_upload_base, list_pendientes_locatario

LOC = "A06_DON_MELCHOR"


def _normalizar_header(h: Any) -> str:
    if h is None:
        return ""
    return str(h).replace("\ufeff", "").strip().lower()


def _primera_fila_headers_csv(path: Path) -> tuple[Any, ...]:
    import csv

    for encoding in ("utf-8", "latin-1"):
        try:
            with path.open(encoding=encoding, errors="replace", newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                header = next(reader, None)
                return tuple(header) if header else ()
        except UnicodeDecodeError:
            continue
    return ()


def _primera_fila_headers_xlsx(path: Path) -> tuple[str, tuple[Any, ...]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        name = ws.title
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = tuple(row) if row else ()
        return name, headers
    finally:
        wb.close()


def _deteccion_tabla_xls(path: Path) -> dict[str, Any]:
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        return {
            "fila_encabezado_0_based": None,
            "indice_columna": None,
            "nombre_columna_elegida": None,
            "encabezado_fila_crudo_muestra": None,
        }
    search_limit = min(ws.nrows, ns._MAX_HEADER_SEARCH_ROWS)
    for i in range(search_limit):
        header = tuple(ws.row_values(i))
        idx = ns._find_fecha_column_index(header)
        if idx is None:
            continue
        found = ns._extract_fechas_bajo_encabezado_xls(ws, wb, i, idx)
        if found:
            nombre_col = str(header[idx]) if idx < len(header) else None
            return {
                "fila_encabezado_0_based": i,
                "indice_columna": idx,
                "nombre_columna_elegida": nombre_col,
                "encabezado_fila_crudo_muestra": [
                    str(header[j]) if j < len(header) else "" for j in range(min(25, len(header)))
                ],
            }
    return {
        "fila_encabezado_0_based": None,
        "indice_columna": None,
        "nombre_columna_elegida": None,
        "encabezado_fila_crudo_muestra": None,
    }


def _deteccion_tabla_xlsx(path: Path) -> dict[str, Any]:
    cap = ns._MAX_HEADER_SEARCH_ROWS + ns._MAX_ROWS_FECHA_COL + 2
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: list[tuple[Any, ...]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(tuple(row) if row else ())
            if i >= cap - 1:
                break
    finally:
        wb.close()
    hit = ns._buscar_fila_encabezado_fecha(rows)
    if hit is None:
        return {
            "fila_encabezado_0_based": None,
            "indice_columna": None,
            "nombre_columna_elegida": None,
            "encabezado_fila_crudo_muestra": None,
        }
    i, idx = hit
    header = rows[i]
    nombre_col = str(header[idx]) if idx < len(header) else None
    return {
        "fila_encabezado_0_based": i,
        "indice_columna": idx,
        "nombre_columna_elegida": nombre_col,
        "encabezado_fila_crudo_muestra": [None if c is None else str(c) for c in header[:25]],
    }


def _diagnostico_un_archivo(path: Path, nombre: str) -> dict[str, Any]:
    ext = path.suffix.lower()
    hoja = None
    headers: tuple[Any, ...] = ()

    if ext == ".xlsx":
        hoja, headers = _primera_fila_headers_xlsx(path)
    elif ext == ".csv":
        headers = _primera_fila_headers_csv(path)
    elif ext == ".xls":
        wb0 = xlrd.open_workbook(str(path))
        hoja = wb0.sheet_names()[0] if wb0.sheet_names() else None
        ws0 = wb0.sheet_by_index(0)
        headers = tuple(ws0.row_values(0)) if ws0.nrows else ()

    headers_norm = [_normalizar_header(h) for h in headers]
    idx_solo_fila1 = ns._find_fecha_column_index(tuple(headers)) if headers else None

    deteccion_tabla: dict[str, Any] = {}
    if ext == ".xlsx":
        deteccion_tabla = _deteccion_tabla_xlsx(path)
    elif ext == ".xls":
        deteccion_tabla = _deteccion_tabla_xls(path)

    fechas_columna: list[str] = []
    try:
        if ext == ".xlsx":
            s = ns._extract_fechas_columna_xlsx(path)
        elif ext == ".xls":
            s = ns._extract_fechas_columna_xls(path)
        elif ext == ".csv":
            s = ns._extract_fechas_columna_csv(path)
        else:
            s = set()
        fechas_columna = sorted(d.isoformat() for d in s)
    except Exception as e:  # noqa: BLE001
        fechas_columna = [f"<error: {e}>"]

    collect = ns.collect_fechas_operacion_desde_archivo(path, nombre)
    collect_iso = sorted(d.isoformat() for d in collect)

    fecha_fb, fuente_fb = ns._fecha_mas_reciente_del_archivo(path, nombre)

    return {
        "archivo": nombre,
        "extension": ext,
        "hoja_activa_xlsx": hoja,
        "fila_1_headers_crudos": [None if h is None else str(h) for h in headers[:40]]
        if headers
        else [],
        "fila_1_headers_normalizados": headers_norm[:40] if headers else [],
        "find_fecha_columna_solo_si_fila_1_es_cabecera": idx_solo_fila1,
        "deteccion_tabla": deteccion_tabla,
        "fechas_extraidas_columna_servicio": fechas_columna,
        "fechas_finales_collect": collect_iso,
        "fallback_si_sin_tabla": {
            "fecha": fecha_fb.isoformat() if fecha_fb else None,
            "fuente": fuente_fb,
        },
    }


def test_diagnostico_fecha_a06_don_melchor() -> None:
    nombres = list_pendientes_locatario(LOC)
    base = get_upload_base()
    d = _dir_locatario_pendientes(base, LOC)

    if not nombres:
        pytest.skip(f"No hay archivos .csv/.xlsx/.xls en pendientes: {d}")

    reporte: dict[str, Any] = {
        "locatario": LOC,
        "directorio": str(d),
        "archivos": [],
    }
    for nombre in nombres:
        path = d / nombre
        if path.is_file():
            reporte["archivos"].append(_diagnostico_un_archivo(path, nombre))

    print("\n=== DIAGNÓSTICO FECHA A06_DON_MELCHOR ===\n")
    print(json.dumps(reporte, ensure_ascii=False, indent=2))
    print("\n=== FIN ===\n")

    assert reporte["archivos"], "lista vacía"
